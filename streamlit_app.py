import os
import shutil
import tempfile
import uuid

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

st.set_page_config(page_title="🚲 E-Bike Tier List & Calculator", layout="wide")

DEFAULT_FILENAME = "ebike tier list blank MK1.xlsx"
ADMIN_PASSWORD = "ebikeadmin1278"
MASTER_HEADER_ROW = 4
MASTER_ROW_START = 5
MASTER_ROW_END = 68
MASTER_COLUMNS = {
    1: "Name:",
    3: "Bike Price ($)",
    6: "Bike Weight (lbs)",
    7: "Rider Weight (lbs)",
    8: "Battery Voltage (V)",
    9: "Battery Amperage (Ah)",
    10: "Peak Power (W)",
    11: "Nominal Power (W)",
    12: "Terrain Type*",
    13: "Driving Style*",
    16: "Has Critical Flaw? (Y/N/C*)",
}
CRITICAL_FLAW_MAP = {
    "No": "N",
    "Yes": "Y",
    "Catastrophic": "C",
}
SORT_PRIORITY = {
    "score": {4: True, 5: False, 6: False, 7: False},
    "speed": {4: False, 5: True, 6: False, 7: False},
    "range": {4: False, 5: False, 6: True, 7: False},
    "price": {4: False, 5: False, 6: False, 7: True},
}

HEADER_LABELS = {
    "Bike name": "Bike name",
    "Price": "Price",
    "Price ($)": "Price",
    "Score (out of 5.01)": "Score",
    "Score": "Score",
    "Unweighted Score": "Score",
}
SUMMARY_COLUMNS = ["Bike name", "Score", "Price"]
SKIP_SCORE_TEXTS = {"Speed (mph)", "Range (mi)", "Price ($)", "Budget Max Price: ($)"}
INVALID_VALUE_PREFIXES = ("#VALUE!",)
RANKED_SCORE_COLUMNS = ["Score (out of 5.01)", "Unweighted Score", "Score"]


def initialize_session_state():
    if "theme" not in st.session_state:
        st.session_state.theme = "light"
    if "profile" not in st.session_state:
        st.session_state.profile = None
    if "pending_submissions" not in st.session_state:
        st.session_state.pending_submissions = []
    if "show_submission_form" not in st.session_state:
        st.session_state.show_submission_form = False
    if "show_admin_password" not in st.session_state:
        st.session_state.show_admin_password = False
    if "admin_authenticated" not in st.session_state:
        st.session_state.admin_authenticated = False
    if "session_id" not in st.session_state:
        st.session_state.session_id = uuid.uuid4().hex
    if "profile_saved" not in st.session_state:
        st.session_state.profile_saved = False
    if "uploaded_filename" not in st.session_state:
        st.session_state.uploaded_filename = None
    if "session_file" not in st.session_state:
        st.session_state.session_file = None


def theme_css():
    if st.session_state.theme == "dark":
        return """
        <style>
            .stApp, .main, .block-container {
                background-color: #0f1720 !important;
                color: #f8fafc !important;
            }
            textarea, input, select, .stButton>button {
                background-color: #1f2937 !important;
                color: #f8fafc !important;
            }
            .stButton>button {
                background-color: #2563eb !important;
                color: white !important;
            }
            label { color: #f8fafc !important; }
        </style>
        """
    return """
        <style>
            .stApp, .main, .block-container {
                background-color: #ffffff !important;
                color: #0f172a !important;
            }
            .stButton>button {
                background-color: #0f62fe !important;
                color: white !important;
            }
            label { color: #0f1720 !important; }
        </style>
        """


def get_session_dir():
    session_dir = os.path.join(tempfile.gettempdir(), "ebike_info_tiering_app")
    os.makedirs(session_dir, exist_ok=True)
    return session_dir


def create_or_get_session_file(uploaded_file):
    if st.session_state.session_file is None or not os.path.exists(st.session_state.session_file):
        st.session_state.session_file = os.path.join(get_session_dir(), f"session_{st.session_state.session_id}.xlsx")

    if uploaded_file is not None:
        current_name = getattr(uploaded_file, "name", None)
        if st.session_state.uploaded_filename != current_name:
            with open(st.session_state.session_file, "wb") as f:
                f.write(uploaded_file.getvalue())
            st.session_state.uploaded_filename = current_name
    elif not os.path.exists(st.session_state.session_file):
        if not os.path.exists(DEFAULT_FILENAME):
            st.error(f"Default workbook not found: {DEFAULT_FILENAME}")
            return None
        shutil.copyfile(DEFAULT_FILENAME, st.session_state.session_file)

    return st.session_state.session_file


def load_master_table(file_path):
    try:
        df = pd.read_excel(
            file_path,
            header=MASTER_HEADER_ROW - 1,
            usecols="A:P",
            nrows=MASTER_ROW_END - MASTER_ROW_START + 1,
            engine="openpyxl",
        )
    except Exception:
        return pd.DataFrame(columns=list(MASTER_COLUMNS.values()))

    rename_map = {
        "Name:": "Name",
        "Rating/5:": "Rating",
        "Bike Price ($)": "Bike Price",
        "Top Speed Potential": "Top Speed Potential",
        "True Throttle Range": "True Throttle Range",
        "Bike Weight (lbs)": "Bike Weight",
        "Rider Weight (lbs)": "Rider Weight",
        "Battery Voltage (V)": "Battery Voltage",
        "Battery Amperage (Ah)": "Battery Amperage",
        "Peak Power (W)": "Peak Power",
        "Nominal Power (W)": "Nominal Power",
        "Terrain Type*": "Terrain Type",
        "Driving Style*": "Driving Style",
        "Battery Size (Wh)": "Battery Size",
        "Battery Burn Rate (Wh/mi)": "Battery Burn Rate",
        "Has Critical Flaw? (Y/N/C*)": "Critical Flaw",
    }
    df = df.rename(columns=rename_map)
    if "Name" in df.columns:
        df = df[df["Name"].notna()].copy()
    return df.reset_index(drop=True)


def get_first_empty_master_row(file_path):
    try:
        workbook = load_workbook(file_path)
        sheet = workbook[workbook.sheetnames[0]]
        for row in range(MASTER_ROW_START, MASTER_ROW_END + 1):
            name_cell = sheet.cell(row=row, column=1).value
            if name_cell is None or str(name_cell).strip() == "":
                workbook.close()
                return row
        workbook.close()
    except Exception:
        pass
    return None


def write_master_row(file_path, row_num, values):
    workbook = load_workbook(file_path)
    sheet = workbook[workbook.sheetnames[0]]
    for col, value in values.items():
        sheet.cell(row=row_num, column=col).value = value
    workbook.save(file_path)
    workbook.close()


def apply_profile_to_file(file_path):
    if st.session_state.profile is None:
        return
    profile = st.session_state.profile
    workbook = load_workbook(file_path)
    sheet = workbook[workbook.sheetnames[0]]
    sheet.cell(row=1, column=8).value = profile["weight"]
    sheet.cell(row=1, column=9).value = profile["target_speed"]
    for row in range(MASTER_ROW_START, MASTER_ROW_END + 1):
        name_value = sheet.cell(row=row, column=1).value
        if name_value is not None and str(name_value).strip() != "":
            sheet.cell(row=row, column=7).value = profile["weight"]
            sheet.cell(row=row, column=12).value = profile["terrain"]
            # Apply driving style derived from user profile
            if "driving_style" in profile:
                sheet.cell(row=row, column=13).value = profile["driving_style"]
    workbook.save(file_path)
    workbook.close()


def _normalize_header(value):
    if value is None:
        return None
    return HEADER_LABELS.get(str(value).strip(), None)


def find_summary_header(sheet):
    for row in sheet.iter_rows(min_row=1, max_row=20, min_col=1, max_col=30):
        label_map = {}
        for cell in row:
            normalized = _normalize_header(cell.value)
            if normalized:
                label_map[normalized] = cell.column
        if "Bike name" in label_map and ("Score" in label_map or "Price" in label_map):
            return row[0].row, label_map
    return None, {}


def _is_blank_value(value):
    if pd.isna(value):
        return True
    text = str(value).strip()
    return text == "" or any(text.startswith(prefix) for prefix in INVALID_VALUE_PREFIXES)


def _is_summary_row(value_map):
    bike_name = value_map.get("Bike name")
    if bike_name is None or pd.isna(bike_name) or str(bike_name).strip() == "":
        return False
    score_value = value_map.get("Score")
    if score_value is None or pd.isna(score_value):
        return False
    score_text = str(score_value).strip()
    if score_text in SKIP_SCORE_TEXTS or score_text.startswith("Note:"):
        return False
    if any(score_text.startswith(prefix) for prefix in INVALID_VALUE_PREFIXES):
        return False
    return True


def _clean_summary_df(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=SUMMARY_COLUMNS)
    df = df.copy()
    for col in SUMMARY_COLUMNS:
        if col not in df.columns:
            df[col] = None
    df = df[SUMMARY_COLUMNS]

    def is_placeholder_row(row):
        bike_name = row["Bike name"]
        score = row["Score"]
        price = row["Price"]
        if _is_blank_value(bike_name) and _is_blank_value(score) and _is_blank_value(price):
            return True
        if bike_name is not None and str(bike_name).strip() == "":
            bike_name = None
        if score is not None:
            score_text = str(score).strip()
            if score_text in SKIP_SCORE_TEXTS or score_text.startswith("Note:") or any(score_text.startswith(prefix) for prefix in INVALID_VALUE_PREFIXES):
                return True
        return False

    cleaned = df[~df.apply(is_placeholder_row, axis=1)]
    return cleaned.reset_index(drop=True)


def extract_summary_table(file):
    try:
        if isinstance(file, (str, os.PathLike)):
            workbook = load_workbook(file, data_only=True)
        else:
            if hasattr(file, "seek"):
                file.seek(0)
            workbook = load_workbook(file, data_only=True)
    except Exception:
        return None
    sheet = workbook[workbook.sheetnames[0]]
    header_row, label_map = find_summary_header(sheet)
    if not label_map:
        return None
    rows = []
    consecutive_empty = 0
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        row_values = {
            label: sheet.cell(row=row_idx, column=label_map.get(label)).value if label_map.get(label) else None
            for label in SUMMARY_COLUMNS
        }
        if all(_is_blank_value(value) for value in row_values.values()):
            consecutive_empty += 1
            if consecutive_empty >= 3:
                break
            continue
        consecutive_empty = 0
        if _is_summary_row(row_values):
            rows.append(row_values)
    if not rows:
        return None
    return _clean_summary_df(pd.DataFrame(rows)[SUMMARY_COLUMNS])


def load_data(file):
    if isinstance(file, (str, os.PathLike)):
        file_path = file
    else:
        file_path = None
    if file_path:
        summary_table = extract_summary_table(file_path)
    else:
        if hasattr(file, "seek"):
            file.seek(0)
        summary_table = extract_summary_table(file)
    if summary_table is not None and not summary_table.empty:
        return summary_table
    if file_path:
        header_row = detect_header_row(file_path)
        df = pd.read_excel(file_path, header=header_row, engine="openpyxl")
    else:
        if hasattr(file, "seek"):
            file.seek(0)
        header_row = detect_header_row(file)
        if hasattr(file, "seek"):
            file.seek(0)
        df = pd.read_excel(file, header=header_row)
    normalized_columns = {}
    for col in df.columns:
        normalized = _normalize_header(col)
        if normalized:
            normalized_columns.setdefault(normalized, []).append(col)
    if normalized_columns:
        result = pd.DataFrame(index=df.index)
        if "Bike name" in normalized_columns:
            result["Bike name"] = df[normalized_columns["Bike name"][0]]
        if "Price" in normalized_columns:
            result["Price"] = df[normalized_columns["Price"][0]]
        score_column = None
        for candidate in RANKED_SCORE_COLUMNS:
            if candidate in df.columns and candidate in normalized_columns.get("Score", []):
                score_column = candidate
                break
        if score_column is None and "Score" in normalized_columns:
            score_column = normalized_columns["Score"][0]
        if score_column is not None:
            result["Score"] = df[score_column]
        if not result.empty:
            return _clean_summary_df(result[SUMMARY_COLUMNS])
    return pd.DataFrame(columns=SUMMARY_COLUMNS)


def detect_header_row(excel_source, marker="Rating/5:", max_rows=20):
    if isinstance(excel_source, (str, os.PathLike)):
        preview = pd.read_excel(excel_source, header=None, nrows=max_rows, engine="openpyxl")
    else:
        if hasattr(excel_source, "seek"):
            excel_source.seek(0)
        preview = pd.read_excel(excel_source, header=None, nrows=max_rows)
    for idx, row in preview.iterrows():
        if marker in row.values:
            return idx
    return 0


def set_sort_priority(file_path, priority):
    if priority not in SORT_PRIORITY:
        return False
    workbook = load_workbook(file_path)
    sheet = workbook[workbook.sheetnames[0]]
    flags = SORT_PRIORITY[priority]
    for row_index, value in flags.items():
        sheet.cell(row=row_index, column=22).value = value
    workbook.save(file_path)
    workbook.close()
    return True


def approve_submission(index, file_path):
    submissions = st.session_state.pending_submissions
    if index < 0 or index >= len(submissions):
        return False
    submission = submissions.pop(index)
    row_num = get_first_empty_master_row(file_path)
    if row_num is None:
        st.error("No empty master row available to approve the bike.")
        return False
    values = {
        1: submission["name"],
        3: submission["price"],
        6: submission["bike_weight"],
        7: st.session_state.profile["weight"],
        8: submission["voltage"],
        9: submission["amperage"],
        10: submission["peak_power"],
        11: submission["nominal_power"],
        12: st.session_state.profile["terrain"],
        13: submission["driving_style"],
        16: CRITICAL_FLAW_MAP.get(submission["critical_flaw"], "N"),
    }
    write_master_row(file_path, row_num, values)
    apply_profile_to_file(file_path)
    st.success(f"Approved '{submission['name']}' and wrote it to master row {row_num}.")
    return True


def deny_submission(index):
    submissions = st.session_state.pending_submissions
    if index < 0 or index >= len(submissions):
        return False
    removed = submissions.pop(index)
    st.warning(f"Denied submission: {removed['name']}")
    return True


def submission_form():
    with st.form("bike_submission", clear_on_submit=True):
        st.subheader("Submit a new e-bike")
        name = st.text_input("Bike name", max_chars=100)
        price = st.number_input("Bike price ($)", min_value=0, value=800)
        bike_weight = st.number_input("Bike weight (lbs)", min_value=0, value=45)
        voltage = st.number_input("Battery voltage (V)", min_value=0, value=48)
        amperage = st.number_input("Battery amperage (Ah)", min_value=0, value=10)
        peak_power = st.number_input("Peak power (W)", min_value=0, value=750)
        nominal_power = st.number_input("Nominal power (W)", min_value=0, value=500)
        driving_style = st.selectbox("Driving style", ["Class 2", "Class 3", "Unlocked"])
        critical_flaw = st.selectbox("Critical flaw level", ["No", "Yes", "Catastrophic"])
        submitted = st.form_submit_button("Send to admin for approval")
    if submitted:
        st.session_state.pending_submissions.append(
            {
                "name": name,
                "price": price,
                "bike_weight": bike_weight,
                "voltage": voltage,
                "amperage": amperage,
                "peak_power": peak_power,
                "nominal_power": nominal_power,
                "driving_style": driving_style,
                "critical_flaw": critical_flaw,
            }
        )
        st.success("Your bike has been submitted for admin approval.")


def admin_panel(file_path):
    st.subheader("Admin panel")
    if not st.session_state.admin_authenticated:
        if st.session_state.show_admin_password:
            password = st.text_input("Admin password", type="password", key="admin_pass")
            if st.button("Unlock admin", key="unlock_admin"):
                if password == ADMIN_PASSWORD:
                    st.session_state.admin_authenticated = True
                    st.success("Admin unlocked.")
                    if hasattr(st, "experimental_rerun"):
                        st.experimental_rerun()
                else:
                    st.error("Incorrect password.")
        return
    submissions = st.session_state.pending_submissions
    left, right = st.columns([2, 3])
    with left:
        st.markdown("#### Pending submissions")
        if not submissions:
            st.info("No pending bike submissions yet.")
        else:
            choices = [f"{idx + 1}. {item['name']}" for idx, item in enumerate(submissions)]
            selected = st.selectbox("Select submission", options=list(range(len(choices))), format_func=lambda i: choices[i])
            submission = submissions[selected]
            st.write("**Suggested bike**")
            st.write(submission)
            if st.button("Approve", key=f"approve_{selected}"):
                approve_submission(selected, file_path)
                if hasattr(st, "experimental_rerun"):
                    st.experimental_rerun()
            if st.button("Deny", key=f"deny_{selected}"):
                deny_submission(selected)
                if hasattr(st, "experimental_rerun"):
                    st.experimental_rerun()
            if st.button("Add test bike (admin only)", key="add_test_admin"):
                add_test_ebike(file_path)
                if hasattr(st, "experimental_rerun"):
                    st.experimental_rerun()
            if st.button("View backend (master table)", key="view_backend"):
                try:
                    master_table = load_master_table(file_path)
                    st.write("#### Backend master table")
                    st.dataframe(master_table)
                except Exception as e:
                    st.error(f"Could not read backend file: {e}")
    with right:
        st.markdown("#### Manual duplicate search")
        query = st.text_input("Search existing master bikes", value=submissions[selected]["name"] if submissions else "")
        master_table = load_master_table(file_path)
        if query:
            matched = master_table[master_table["Name"].astype(str).str.contains(query, case=False, na=False)]
        else:
            matched = master_table.head(10)
        st.write(f"Master records matching '{query}'" if query else "Recent master entries")
        st.dataframe(matched[["Name", "Bike Price", "Top Speed Potential", "True Throttle Range"]].head(20))


def add_test_ebike(file_path):
    row_num = get_first_empty_master_row(file_path)
    if row_num is None:
        st.error("❌ No available rows in the master table. All 64 bike slots are filled.")
        return False
    try:
        values = {
            1: "Test Budget eBike",
            3: 800,
            6: 45,
            7: st.session_state.profile["weight"] if st.session_state.profile else 180,
            8: 48,
            9: 10,
            10: 750,
            11: 500,
            12: st.session_state.profile["terrain"] if st.session_state.profile else "Urban",
            13: st.session_state.profile.get("driving_style", "Unlocked") if st.session_state.profile else "Unlocked",
            16: "N",
        }
        write_master_row(file_path, row_num, values)
        apply_profile_to_file(file_path)
        st.success(f"✅ Test eBike added to master row {row_num}.")
        return True
    except Exception as e:
        st.error(f"❌ Error adding test eBike: {e}")
        return False


def render_leaderboard(file_path):
    st.subheader("🏆 Top 10 E-Bikes Leaderboard")
    leaderboard = load_data(file_path)
    if leaderboard is not None and not leaderboard.empty:
        display_df = leaderboard.head(10).reset_index(drop=True)
        display_df.index = display_df.index + 1
        st.dataframe(display_df)
    else:
        st.warning("The workbook contains a summary table template, but no bike rows are populated yet.")
    st.write("---")
    cols = st.columns(4)
    if cols[0].button("Sort by Score"):
        if set_sort_priority(file_path, "score"):
            if hasattr(st, "experimental_rerun"):
                st.experimental_rerun()
    if cols[1].button("Sort by Speed"):
        if set_sort_priority(file_path, "speed"):
            if hasattr(st, "experimental_rerun"):
                st.experimental_rerun()
    if cols[2].button("Sort by Range"):
        if set_sort_priority(file_path, "range"):
            if hasattr(st, "experimental_rerun"):
                st.experimental_rerun()
    if cols[3].button("Sort by Price"):
        if set_sort_priority(file_path, "price"):
            if hasattr(st, "experimental_rerun"):
                st.experimental_rerun()


def profile_form():
    st.header("Welcome — set your rider profile")
    with st.form("profile_form"):
        weight = st.number_input("Input user weight (lbs)", min_value=1, max_value=500, value=180)
        terrain = st.selectbox("Select riding terrain", ["Urban", "Suburban", "Trail"])
        target_speed = st.number_input("Target speed (mph)", min_value=1, max_value=120, value=20)
        st.markdown("_Driving style will be inferred from target speed (<=20 → Class 2, <=28 → Class 3, >28 → Unlocked)_")
        submit = st.form_submit_button("Save profile")
    if submit:
        # derive driving style from target speed
        if target_speed <= 20:
            driving_style = "Class 2"
        elif target_speed <= 28:
            driving_style = "Class 3"
        else:
            driving_style = "Unlocked"

        st.session_state.profile = {
            "weight": weight,
            "terrain": terrain,
            "target_speed": target_speed,
            "driving_style": driving_style,
        }
        st.session_state.profile_saved = True
        session_file = create_or_get_session_file(None)
        if session_file:
            apply_profile_to_file(session_file)


def main():
    initialize_session_state()
    st.markdown(theme_css(), unsafe_allow_html=True)

    # Use application default workbook copy (upload removed in deployed app)
    active_file = create_or_get_session_file(None)
    if active_file is None:
        return

    if st.session_state.profile is None:
        profile_form()
        if st.session_state.profile_saved:
            st.session_state.profile_saved = False
            if hasattr(st, "experimental_rerun"):
                st.experimental_rerun()
            else:
                # Streamlit runtime doesn't expose experimental_rerun in this environment.
                # Rely on normal rerun behaviour triggered by widgets; continue.
                pass
        return

    apply_profile_to_file(active_file)

    top_cols = st.columns([2.6, 2.6, 1])
    with top_cols[0]:
        if st.button("Submit a Bike", key="submit_bike", use_container_width=True):
            st.session_state.show_submission_form = True
    with top_cols[1]:
        if st.button("Admin Panel", key="admin_panel", use_container_width=True):
            st.session_state.show_admin_password = True
    with top_cols[2]:
        theme_choice = st.selectbox("Theme", ["Light", "Dark"], index=0 if st.session_state.theme == "light" else 1)
        st.session_state.theme = "dark" if theme_choice == "Dark" else "light"

    if st.session_state.show_submission_form:
        submission_form()

    admin_panel(active_file)
    st.write("---")
    render_leaderboard(active_file)


if __name__ == "__main__":
    main()
